from __future__ import annotations

import csv
import io
from collections import defaultdict
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb_workbook


REQUIRED_COLUMNS = [
    "Practice Area",
    "Parent Account Name",
    "FY 26 (SL)",
    "Target 2026",
    "SLS",
]


def _get_column(col_map: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in col_map:
            return col_map[name]
    return None


def _to_number(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _dollars_to_millions(value: float) -> float:
    return value / 1000


def _money_label(value: float) -> str:
    return f"${_dollars_to_millions(value):,.1f}M"


def _gap_status(value: float) -> str:
    gap_in_millions = _dollars_to_millions(value)
    if gap_in_millions > 0.05:
        return "behind"
    if gap_in_millions < -0.05:
        return "ahead"
    return "on-track"


def analyze_forecast_rows(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    sls_name: str,
) -> dict[str, Any]:
    name = str(sls_name or "").strip()
    if not name:
        raise ValueError("SLS name is required.")

    missing = [column for column in REQUIRED_COLUMNS if column not in col_map]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    c_sls = col_map["SLS"]
    c_practice = col_map["Practice Area"]
    c_account = col_map["Parent Account Name"]
    c_source = _get_column(col_map, "P&L Source", "PLSource")
    c_header = _get_column(col_map, "P&L Header", "PLHeader")
    c_forecast = col_map["FY 26 (SL)"]
    c_target = col_map["Target 2026"]

    if c_source is None:
        raise ValueError("Missing columns: P&L Source or PLSource")
    if c_header is None:
        raise ValueError("Missing columns: P&L Header or PLHeader")

    forecast: dict[str, float] = defaultdict(float)
    target: dict[str, float] = defaultdict(float)
    matched_sls_names: set[str] = set()
    normalized_name = name.lower()

    for row in data_rows:
        sls = _cell(row, c_sls)
        if not sls:
            continue

        sls_text = str(sls).strip()
        if normalized_name not in sls_text.lower():
            continue

        header = str(_cell(row, c_header) or "").strip()
        if header != "Net Revenue":
            continue

        source = str(_cell(row, c_source) or "").strip()
        account = str(_cell(row, c_account) or "").strip()
        practice = str(_cell(row, c_practice) or "").strip()
        key = f"{account}__{practice}"
        matched_sls_names.add(sls_text)

        if source == "IC/Forecasted":
            forecast[key] += _to_number(_cell(row, c_forecast))
        elif source == "Budget":
            target[key] += _to_number(_cell(row, c_target))

    rows = []
    for key in sorted(set(forecast) | set(target)):
        account, practice = key.split("__", 1)
        forecast_amount = forecast[key]
        target_amount = target[key]
        gap_amount = target_amount - forecast_amount

        if abs(_dollars_to_millions(forecast_amount)) < 0.05 and abs(_dollars_to_millions(target_amount)) < 0.05:
            continue

        rows.append(
            {
                "account": account,
                "practice": practice,
                "forecast": forecast_amount,
                "target": target_amount,
                "gap": gap_amount,
                "labels": {
                    "forecast": _money_label(forecast_amount),
                    "target": _money_label(target_amount),
                    "gap": _money_label(gap_amount),
                },
                "status": _gap_status(gap_amount),
            }
        )

    rows.sort(key=lambda row: (row["account"], row["practice"]))

    account_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        account_name = row["account"]
        account = account_map.setdefault(
            account_name,
            {
                "account": account_name,
                "forecast": 0.0,
                "target": 0.0,
                "gap": 0.0,
                "practices": [],
            },
        )
        account["forecast"] += row["forecast"]
        account["target"] += row["target"]
        account["gap"] += row["gap"]
        account["practices"].append(row)

    accounts = []
    for account in account_map.values():
        accounts.append(
            {
                **account,
                "labels": {
                    "forecast": _money_label(account["forecast"]),
                    "target": _money_label(account["target"]),
                    "gap": _money_label(account["gap"]),
                },
                "status": _gap_status(account["gap"]),
            }
        )

    totals = {
        "forecast": sum(row["forecast"] for row in rows),
        "target": sum(row["target"] for row in rows),
    }
    totals["gap"] = totals["target"] - totals["forecast"]

    return {
        "query": name,
        "matchedSlsNames": sorted(matched_sls_names),
        "metrics": {
            **totals,
            "accounts": len(accounts),
            "rows": len(rows),
            "labels": {
                "forecast": _money_label(totals["forecast"]),
                "target": _money_label(totals["target"]),
                "gap": _money_label(totals["gap"]),
            },
            "status": _gap_status(totals["gap"]),
        },
        "accounts": accounts,
        "rows": rows,
    }


def parse_workbook(file_bytes: bytes, filename: str, sheet_name: str = "Data") -> tuple[list[str], list[list[Any]], dict[str, int]]:
    values = _read_sheet(file_bytes, filename, sheet_name)
    if len(values) < 2:
        raise ValueError(f"{sheet_name} sheet appears empty.")

    headers = [str(header).strip() if header is not None and str(header).strip() else f"Column {index + 1}" for index, header in enumerate(values[0])]
    col_map = {header: index for index, header in enumerate(headers) if header}
    return headers, values[1:], col_map


def analyze_workbook(file_bytes: bytes, filename: str, sls_name: str, sheet_name: str = "Data") -> dict[str, Any]:
    _headers, rows, col_map = parse_workbook(file_bytes, filename, sheet_name)
    return analyze_forecast_rows(rows, col_map, sls_name)


def result_to_csv(result: dict[str, Any]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(["Account", "Practice", "Forecast_$K", "Target_$K", "Gap_$K"])

    for row in result.get("rows", []):
        writer.writerow(
            [
                row["account"],
                row["practice"],
                f'{row["forecast"]:.2f}',
                f'{row["target"]:.2f}',
                f'{row["gap"]:.2f}',
            ]
        )

    return output.getvalue()


def _read_sheet(file_bytes: bytes, filename: str, sheet_name: str) -> list[list[Any]]:
    extension = Path(filename or "").suffix.lower()
    if extension == ".xlsb":
        return _read_xlsb(file_bytes, sheet_name)
    if extension in {".xlsx", ".xlsm", ""}:
        return _read_xlsx(file_bytes, sheet_name)
    raise ValueError("Unsupported workbook type. Upload .xlsx, .xlsm, or .xlsb.")


def _read_xlsx(file_bytes: bytes, sheet_name: str) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found in workbook.')

    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _read_xlsb(file_bytes: bytes, sheet_name: str) -> list[list[Any]]:
    with NamedTemporaryFile(suffix=".xlsb") as tmp:
        tmp.write(file_bytes)
        tmp.flush()

        with open_xlsb_workbook(tmp.name) as workbook:
            if sheet_name not in workbook.sheets:
                raise ValueError(f'Sheet "{sheet_name}" not found in workbook.')

            with workbook.get_sheet(sheet_name) as sheet:
                return [[cell.v for cell in row] for row in sheet.rows()]


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None
