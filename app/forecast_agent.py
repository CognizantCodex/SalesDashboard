from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from pyxlsb import Workbook as XlsbWorkbook


PRACTICE_COLUMNS = ("Practice Area", "Practice")
REVENUE_ACCOUNT_COLUMNS = ("Parent Account Name", "Financial Ultimate Parent Account", "Account Name")
FORECAST_AMOUNT_COLUMNS = ("FY 26 (SL)", "CY $", "Current Year Revenue (converted)", "CY REVENUE $")
TARGET_AMOUNT_COLUMNS = ("Target 2026", "Revenue Target", "Target Revenue", "CY Target")


PIPELINE_REQUIRED_COLUMNS = [
    "Grouped Sales Stage",
    "EDC Year",
]

PIPELINE_STAGES = {"Qualified", "Un-Qualified"}
PIPELINE_EXCLUDED_SUB_STATUSES = {"negotiation"}
WON_LOST_STAGES = {"Won", "Lost"}
PENDING_VALIDATION_STATUS = "Pending Validation"
SLSM_FORECAST_SHEET = "SL_Forecast -2026"
DEAL_TYPE_COLUMNS = ("Grouped Deal Type", "Group Deal Type", "Deal Type")
TARGETS_SHEET = "SLM-SLS-Pivot"
TARGET_TOTAL_LABEL = "Grand Total"


def _get_column(col_map: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in col_map:
            return col_map[name]

    normalized_map = {
        _normalize_column_name(column): index
        for column, index in col_map.items()
    }
    for name in names:
        normalized_name = _normalize_column_name(name)
        if normalized_name in normalized_map:
            return normalized_map[normalized_name]
    return None


def _get_column_name(col_map: dict[str, int], *names: str) -> str | None:
    for name in names:
        if name in col_map:
            return name

    normalized_map = {
        _normalize_column_name(column): column
        for column in col_map
    }
    for name in names:
        normalized_name = _normalize_column_name(name)
        if normalized_name in normalized_map:
            return normalized_map[normalized_name]
    return None


def _normalize_column_name(value: str) -> str:
    text = str(value or "").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _get_deal_type_column(col_map: dict[str, int]) -> int | None:
    return _get_column(col_map, *DEAL_TYPE_COLUMNS)


def _person_column_candidates(person_column: str) -> tuple[str, ...]:
    if person_column == "SLSM":
        return ("SLSM", "SLSM Name", "SLS Manager")
    return (person_column,)


def _get_person_column(col_map: dict[str, int], person_column: str) -> int | None:
    return _get_column(col_map, *_person_column_candidates(person_column))


def _person_missing_label(person_column: str) -> str:
    candidates = _person_column_candidates(person_column)
    if len(candidates) == 1:
        return candidates[0]
    return " or ".join(candidates)


def unique_person_values(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    person_column: str,
) -> list[str]:
    person_column_index = _get_person_column(col_map, person_column)
    if person_column_index is None:
        raise ValueError(f"Missing columns: {_person_missing_label(person_column)}")

    values = {
        str(_cell(row, person_column_index) or "").strip()
        for row in data_rows
        if str(_cell(row, person_column_index) or "").strip()
    }
    return sorted(values, key=lambda value: value.lower())


def _to_number(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _dollars_to_millions(value: float) -> float:
    return value / 1000


def _money_label(value: float) -> str:
    return f"${_dollars_to_millions(value):,.1f}M"


def _revenue_amount(row: list[Any], column: int | None, column_name: str | None) -> float:
    if column is None:
        return 0.0

    amount = _to_number(_cell(row, column))
    if column_name in {"CY $", "Current Year Revenue (converted)", "CY REVENUE $", "Revenue Target", "Target Revenue", "CY Target"}:
        return amount / 1000
    return amount


def _pipeline_money_label(value: float) -> str:
    return f"${value / 1_000_000:,.1f}M"


def _target_value_label(header: str, value: float) -> str:
    if "OM %" in header.upper():
        return f"{value:.1%}"
    return f"${value / 1_000_000:,.1f}M"


def _gap_status(value: float) -> str:
    gap_in_millions = _dollars_to_millions(value)
    if gap_in_millions > 0.05:
        return "behind"
    if gap_in_millions < -0.05:
        return "ahead"
    return "on-track"


def _is_active_row(row: list[Any], col_map: dict[str, int]) -> bool:
    active_column = _get_column(col_map, "Is Active?")
    if active_column is None:
        return True

    value = str(_cell(row, active_column) or "").strip().lower()
    return value in {"1", "1.0", "true", "yes"}


def _is_master_contract_blank(row: list[Any], col_map: dict[str, int]) -> bool:
    master_contract_column = _get_column(col_map, "Master Contract Opportunity")
    if master_contract_column is None:
        return True

    value = str(_cell(row, master_contract_column) or "").strip()
    return value in {"", "(blank)"}


def analyze_forecast_rows(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    sls_name: str,
    person_column: str = "SLS",
) -> dict[str, Any]:
    name = str(sls_name or "").strip()
    if not name:
        raise ValueError(f"{person_column} name is required.")

    person_column_index = _get_person_column(col_map, person_column)
    practice_column = _get_column(col_map, *PRACTICE_COLUMNS)
    account_column = _get_column(col_map, *REVENUE_ACCOUNT_COLUMNS)
    forecast_column = _get_column(col_map, *FORECAST_AMOUNT_COLUMNS)
    target_column = _get_column(col_map, *TARGET_AMOUNT_COLUMNS)
    forecast_column_name = _get_column_name(col_map, *FORECAST_AMOUNT_COLUMNS)
    target_column_name = _get_column_name(col_map, *TARGET_AMOUNT_COLUMNS)
    missing = []
    if person_column_index is None:
        missing.append(_person_missing_label(person_column))
    if practice_column is None:
        missing.append("Practice Area or Practice")
    if account_column is None:
        missing.append("Parent Account Name, Financial Ultimate Parent Account, or Account Name")
    if forecast_column is None:
        missing.append("FY 26 (SL), CY $, Current Year Revenue (converted), or CY REVENUE $")
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    c_source = _get_column(col_map, "P&L Source", "PLSource")
    c_header = _get_column(col_map, "P&L Header", "PLHeader")
    uses_source_rows = c_source is not None and c_header is not None

    # A revenue account can use a combined SLS assignment for budget rows but
    # an individual SLS assignment for forecast rows. Build the matching
    # participants per account first, so a partial search such as "som" can
    # include the associated forecast owner without including that owner's
    # unrelated accounts.
    account_participants: dict[str, set[str]] = defaultdict(set)
    for row in data_rows:
        person_text = str(_cell(row, person_column_index) or "").strip()
        account = str(_cell(row, account_column) or "").strip()
        if account and person_text and _matches_name_permutation(person_text, name):
            account_participants[account].update(_split_combined_name(person_text))

    forecast: dict[str, float] = defaultdict(float)
    target: dict[str, float] = defaultdict(float)
    matched_sls_names: set[str] = set()
    for row in data_rows:
        person_value = _cell(row, person_column_index)
        if not person_value:
            continue

        person_text = str(person_value).strip()
        account = str(_cell(row, account_column) or "").strip()
        participant_match = any(
            _matches_name_permutation(person_text, participant)
            for participant in account_participants.get(account, set())
        )
        if not _matches_name_permutation(person_text, name) and not participant_match:
            continue

        practice = str(_cell(row, practice_column) or "").strip()
        key = f"{account}__{practice}"
        matched_sls_names.add(person_text)

        if uses_source_rows:
            header = str(_cell(row, c_header) or "").strip()
            if header != "Net Revenue":
                continue

            source = str(_cell(row, c_source) or "").strip()
            if source == "IC/Forecasted":
                forecast[key] += _revenue_amount(row, forecast_column, forecast_column_name)
            elif source == "Budget" and target_column is not None:
                target[key] += _revenue_amount(row, target_column, target_column_name)
        else:
            forecast[key] += _revenue_amount(row, forecast_column, forecast_column_name)
            if target_column is not None:
                target[key] += _revenue_amount(row, target_column, target_column_name)

    rows = []
    for key in sorted(set(forecast) | set(target)):
        account, practice = key.split("__", 1)
        forecast_amount = forecast[key]
        target_amount = target[key]
        gap_amount = target_amount - forecast_amount

        if abs(_dollars_to_millions(forecast_amount)) < 0.05 and abs(_dollars_to_millions(target_amount)) < 0.05:
            continue

        rows.append(
            {
                "account": account,
                "practice": practice,
                "forecast": forecast_amount,
                "target": target_amount,
                "gap": gap_amount,
                "labels": {
                    "forecast": _money_label(forecast_amount),
                    "target": _money_label(target_amount),
                    "gap": _money_label(gap_amount),
                },
                "status": _gap_status(gap_amount),
            }
        )

    rows.sort(key=lambda row: (row["account"], row["practice"]))

    account_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        account_name = row["account"]
        account = account_map.setdefault(
            account_name,
            {
                "account": account_name,
                "forecast": 0.0,
                "target": 0.0,
                "gap": 0.0,
                "practices": [],
            },
        )
        account["forecast"] += row["forecast"]
        account["target"] += row["target"]
        account["gap"] += row["gap"]
        account["practices"].append(row)

    accounts = []
    for account in account_map.values():
        accounts.append(
            {
                **account,
                "labels": {
                    "forecast": _money_label(account["forecast"]),
                    "target": _money_label(account["target"]),
                    "gap": _money_label(account["gap"]),
                },
                "status": _gap_status(account["gap"]),
            }
        )

    totals = {
        "forecast": sum(row["forecast"] for row in rows),
        "target": sum(row["target"] for row in rows),
    }
    totals["gap"] = totals["target"] - totals["forecast"]
    return {
        "query": name,
        "matchedSlsNames": sorted(matched_sls_names),
        "metrics": {
            **totals,
            "accounts": len(accounts),
            "rows": len(rows),
            "labels": {
                "forecast": _money_label(totals["forecast"]),
                "target": _money_label(totals["target"]),
                "gap": _money_label(totals["gap"]),
            },
            "status": _gap_status(totals["gap"]),
        },
        "accounts": accounts,
        "rows": rows,
    }


def analyze_pipeline_rows(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    sls_name: str,
    current_year: int | None = None,
    person_column: str = "SLS",
) -> dict[str, Any]:
    name = str(sls_name or "").strip()
    if not name:
        raise ValueError(f"{person_column} name is required.")

    missing = [column for column in PIPELINE_REQUIRED_COLUMNS if column not in col_map]
    person_column_index = _get_person_column(col_map, person_column)
    if person_column_index is None:
        missing.append(_person_missing_label(person_column))
    amount_column = _get_column(
        col_map,
        "Net TCV Share",
        "Net TCV Share (converted)",
        "CY $",
        "Current Year Revenue (converted)",
        "CY REVENUE $",
    )
    account_column = _get_column(col_map, "Financial Ultimate Parent Account", "Account Name")
    practice_column = _get_column(col_map, "Practice Area", "Practice")
    deal_type_column = _get_deal_type_column(col_map)
    if amount_column is None:
        missing.append("Net TCV Share or CY $")
    if account_column is None:
        missing.append("Financial Ultimate Parent Account or Account Name")
    if practice_column is None:
        missing.append("Practice Area or Practice")
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    c_stage = col_map["Grouped Sales Stage"]
    c_year = col_map["EDC Year"]
    sub_status_column = _get_column(col_map, "Sub-Status")
    year = current_year or date.today().year

    totals = {
        "pipeline": 0.0,
        "qualified": 0.0,
        "unqualified": 0.0,
    }
    accounts: set[str] = set()
    matched_sls_names: set[str] = set()
    account_practices: dict[str, dict[str, Any]] = {}
    rows_saved = 0

    for row in data_rows:
        stage = str(_cell(row, c_stage) or "").strip()
        if stage not in PIPELINE_STAGES:
            continue

        if sub_status_column is not None:
            sub_status = str(_cell(row, sub_status_column) or "").strip().lower()
            if sub_status in PIPELINE_EXCLUDED_SUB_STATUSES:
                continue

        if _year_value(_cell(row, c_year)) != year:
            continue

        sls_text = str(_cell(row, person_column_index) or "").strip()
        if not _matches_name_permutation(sls_text, name):
            continue

        amount = _to_number(_cell(row, amount_column))
        account = str(_cell(row, account_column) or "").strip()
        practice = str(_cell(row, practice_column) or "").strip()
        deal_type = str(_cell(row, deal_type_column) or "").strip() if deal_type_column is not None else ""
        deal_type = deal_type or "Unspecified"
        if account:
            accounts.add(account)
        matched_sls_names.add(sls_text)
        rows_saved += 1

        key = f"{account}__{practice}__{deal_type}"
        detail = account_practices.setdefault(
            key,
            {
                "account": account,
                "practice": practice,
                "dealType": deal_type,
                "pipeline": 0.0,
                "qualified": 0.0,
                "unqualified": 0.0,
                "rows": 0,
            },
        )
        detail["pipeline"] += amount
        detail["rows"] += 1

        totals["pipeline"] += amount
        if stage == "Qualified":
            totals["qualified"] += amount
            detail["qualified"] += amount
        elif stage == "Un-Qualified":
            totals["unqualified"] += amount
            detail["unqualified"] += amount

    detail_rows = []
    for detail in sorted(account_practices.values(), key=lambda item: (item["account"], item["practice"])):
        detail_rows.append(
            {
                **detail,
                "labels": {
                    "pipeline": _pipeline_money_label(detail["pipeline"]),
                    "qualified": _pipeline_money_label(detail["qualified"]),
                    "unqualified": _pipeline_money_label(detail["unqualified"]),
                },
            }
        )

    account_map: dict[str, dict[str, Any]] = {}
    for detail in detail_rows:
        account_name = detail["account"]
        account = account_map.setdefault(
            account_name,
            {
                "account": account_name,
                "pipeline": 0.0,
                "qualified": 0.0,
                "unqualified": 0.0,
                "rows": 0,
                "practices": [],
            },
        )
        account["pipeline"] += detail["pipeline"]
        account["qualified"] += detail["qualified"]
        account["unqualified"] += detail["unqualified"]
        account["rows"] += detail["rows"]
        account["practices"].append(detail)

    account_rows = [
        {
            **account,
            "labels": {
                "pipeline": _pipeline_money_label(account["pipeline"]),
                "qualified": _pipeline_money_label(account["qualified"]),
                "unqualified": _pipeline_money_label(account["unqualified"]),
            },
        }
        for account in account_map.values()
    ]

    return {
        "available": True,
        "query": name,
        "year": year,
        "matchedSlsNames": sorted(matched_sls_names),
        "metrics": {
            **totals,
            "accounts": len(accounts),
            "rows": rows_saved,
            "labels": {
                "pipeline": _pipeline_money_label(totals["pipeline"]),
                "qualified": _pipeline_money_label(totals["qualified"]),
                "unqualified": _pipeline_money_label(totals["unqualified"]),
            },
        },
        "accounts": account_rows,
        "rows": detail_rows,
    }


def analyze_won_lost_rows(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    sls_name: str,
    current_year: int | None = None,
    person_column: str = "SLS",
) -> dict[str, Any]:
    name = str(sls_name or "").strip()
    if not name:
        raise ValueError(f"{person_column} name is required.")

    missing = [column for column in PIPELINE_REQUIRED_COLUMNS if column not in col_map]
    person_column_index = _get_person_column(col_map, person_column)
    if person_column_index is None:
        missing.append(_person_missing_label(person_column))
    amount_column = _get_column(col_map, "Net TCV Share", "Net TCV Share (converted)")
    account_column = _get_column(col_map, "Financial Ultimate Parent Account", "Account Name")
    practice_column = _get_column(col_map, "Practice Area", "Practice")
    deal_type_column = _get_deal_type_column(col_map)
    if amount_column is None:
        missing.append("Net TCV Share")
    if account_column is None:
        missing.append("Financial Ultimate Parent Account or Account Name")
    if practice_column is None:
        missing.append("Practice Area or Practice")
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    c_stage = col_map["Grouped Sales Stage"]
    c_year = _get_column(col_map, "Year Closed") or col_map["EDC Year"]
    year = current_year or date.today().year

    totals = {"won": 0.0, "lost": 0.0}
    matched_sls_names: set[str] = set()
    account_practices: dict[str, dict[str, Any]] = {}
    rows_saved = 0

    for row in data_rows:
        stage = str(_cell(row, c_stage) or "").strip()
        normalized_stage = stage.lower()
        if normalized_stage not in {"won", "lost"}:
            continue

        if not _is_active_row(row, col_map) or not _is_master_contract_blank(row, col_map):
            continue

        if _year_value(_cell(row, c_year)) != year:
            continue

        sls_text = str(_cell(row, person_column_index) or "").strip()
        if not _matches_name_permutation(sls_text, name):
            continue

        amount = _to_number(_cell(row, amount_column))
        account = str(_cell(row, account_column) or "").strip()
        practice = str(_cell(row, practice_column) or "").strip()
        deal_type = str(_cell(row, deal_type_column) or "").strip() if deal_type_column is not None else ""
        deal_type = deal_type or "Unspecified"
        detail = account_practices.setdefault(
            f"{account}__{practice}__{deal_type}",
            {
                "account": account,
                "practice": practice,
                "dealType": deal_type,
                "won": 0.0,
                "lost": 0.0,
                "total": 0.0,
                "rows": 0,
            },
        )

        totals[normalized_stage] += amount
        detail[normalized_stage] += amount
        detail["total"] = detail["won"]
        detail["rows"] += 1
        matched_sls_names.add(sls_text)
        rows_saved += 1

    total = totals["won"]
    detail_rows = _build_tcv_account_rows(account_practices, "won", "total")

    return {
        "available": True,
        "query": name,
        "year": year,
        "matchedSlsNames": sorted(matched_sls_names),
        "metrics": {
            **totals,
            "total": total,
            "rows": rows_saved,
            "labels": {
                "total": _pipeline_money_label(total),
                "won": _pipeline_money_label(totals["won"]),
                "lost": _pipeline_money_label(totals["lost"]),
            },
        },
        "accounts": detail_rows["accounts"],
        "rows": detail_rows["rows"],
    }


def analyze_pending_validation_rows(
    data_rows: list[list[Any]],
    col_map: dict[str, int],
    sls_name: str,
    current_year: int | None = None,
    person_column: str = "SLS",
) -> dict[str, Any]:
    name = str(sls_name or "").strip()
    if not name:
        raise ValueError(f"{person_column} name is required.")

    required_columns = ["Sub-Status", "EDC Year"]
    missing = [column for column in required_columns if column not in col_map]
    person_column_index = _get_person_column(col_map, person_column)
    if person_column_index is None:
        missing.append(_person_missing_label(person_column))
    amount_column = _get_column(col_map, "Net TCV Share", "Net TCV Share (converted)")
    account_column = _get_column(col_map, "Financial Ultimate Parent Account", "Account Name")
    practice_column = _get_column(col_map, "Practice Area", "Practice")
    deal_type_column = _get_deal_type_column(col_map)
    if amount_column is None:
        missing.append("Net TCV Share")
    if account_column is None:
        missing.append("Financial Ultimate Parent Account or Account Name")
    if practice_column is None:
        missing.append("Practice Area or Practice")
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    c_status = col_map["Sub-Status"]
    c_year = col_map["EDC Year"]
    year = current_year or date.today().year

    total = 0.0
    matched_sls_names: set[str] = set()
    account_practices: dict[str, dict[str, Any]] = {}
    rows_saved = 0

    for row in data_rows:
        status = str(_cell(row, c_status) or "").strip()
        if status.lower() != PENDING_VALIDATION_STATUS.lower():
            continue

        if not _is_active_row(row, col_map) or not _is_master_contract_blank(row, col_map):
            continue

        if _year_value(_cell(row, c_year)) != year:
            continue

        sls_text = str(_cell(row, person_column_index) or "").strip()
        if not _matches_name_permutation(sls_text, name):
            continue

        amount = _to_number(_cell(row, amount_column))
        account = str(_cell(row, account_column) or "").strip()
        practice = str(_cell(row, practice_column) or "").strip()
        deal_type = str(_cell(row, deal_type_column) or "").strip() if deal_type_column is not None else ""
        deal_type = deal_type or "Unspecified"
        detail = account_practices.setdefault(
            f"{account}__{practice}__{deal_type}",
            {
                "account": account,
                "practice": practice,
                "dealType": deal_type,
                "pendingValidation": 0.0,
                "rows": 0,
            },
        )
        total += amount
        detail["pendingValidation"] += amount
        detail["rows"] += 1
        matched_sls_names.add(sls_text)
        rows_saved += 1

    detail_rows = _build_tcv_account_rows(account_practices, "pendingValidation")

    return {
        "available": True,
        "query": name,
        "year": year,
        "matchedSlsNames": sorted(matched_sls_names),
        "metrics": {
            "pendingValidation": total,
            "rows": rows_saved,
            "labels": {
                "pendingValidation": _pipeline_money_label(total),
            },
        },
        "accounts": detail_rows["accounts"],
        "rows": detail_rows["rows"],
    }



def _build_tcv_account_rows(account_practices: dict[str, dict[str, Any]], *amount_keys: str) -> dict[str, list[dict[str, Any]]]:
    detail_rows = []
    for detail in sorted(account_practices.values(), key=lambda item: (item["account"], item["practice"])):
        labels = {key: _pipeline_money_label(detail.get(key, 0.0)) for key in amount_keys}
        detail_rows.append({**detail, "labels": labels})

    account_map: dict[str, dict[str, Any]] = {}
    for detail in detail_rows:
        account_name = detail["account"]
        account = account_map.setdefault(
            account_name,
            {
                "account": account_name,
                "rows": 0,
                "practices": [],
                **{key: 0.0 for key in amount_keys},
            },
        )
        account["rows"] += detail.get("rows", 0)
        for key in amount_keys:
            account[key] += detail.get(key, 0.0)
        account["practices"].append(detail)

    account_rows = []
    for account in account_map.values():
        labels = {key: _pipeline_money_label(account.get(key, 0.0)) for key in amount_keys}
        account_rows.append({**account, "labels": labels})

    return {"accounts": account_rows, "rows": detail_rows}


def _matches_name_permutation(value: Any, query: str) -> bool:
    value_text = str(value or "").strip()
    if not value_text:
        return False

    normalized_value = _normalize_name(value_text)
    normalized_query = _normalize_name(query)
    if normalized_query and normalized_query in normalized_value:
        return True

    value_candidates = _split_combined_name(value_text)
    query_candidates = _split_combined_name(query)
    return any(
        all(token in _normalize_name(value_candidate) for token in _name_tokens(query_candidate))
        for value_candidate in value_candidates
        for query_candidate in query_candidates
        if _name_tokens(query_candidate)
    )


def _split_combined_name(value: str) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []

    # Forecast workbooks may assign the forecast to an individual SLS while
    # assigning budget to a combined SLS label, e.g. "Ali, Afzal - Das,Somnath".
    # Treat the individual names as candidates so both rows remain in the same
    # revenue result.
    parts = [part.strip() for part in re.split(r"\s*(?:\+|\s+-\s+)\s*", text) if part.strip()]
    return [text, *parts]


def _name_tokens(value: str) -> list[str]:
    return [token for token in re.split(r"[^a-z0-9]+", str(value or "").lower()) if token]


def _normalize_name(value: str) -> str:
    return " ".join(_name_tokens(value))


def _year_value(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        match = re.search(r"(20\d{2})", str(value))
        return int(match.group(1)) if match else None


def parse_workbook(file_bytes: bytes, filename: str, sheet_name: str = "Data") -> tuple[list[str], list[list[Any]], dict[str, int]]:
    values = _read_sheet(file_bytes, filename, sheet_name)
    if len(values) < 2:
        raise ValueError(f"{sheet_name} sheet appears empty.")

    headers = [str(header).strip() if header is not None and str(header).strip() else f"Column {index + 1}" for index, header in enumerate(values[0])]
    col_map = {header: index for index, header in enumerate(headers) if header}
    return headers, values[1:], col_map


def parse_target_pivot(file_bytes: bytes, filename: str, sheet_name: str = TARGETS_SHEET) -> dict[str, Any]:
    values = _read_sheet(file_bytes, filename, sheet_name)
    header_index = next(
        (
            index
            for index, row in enumerate(values)
            if str(_cell(row, 0) or "").strip().lower() == "row labels"
        ),
        None,
    )
    if header_index is None:
        raise ValueError('Missing "Row Labels" column in SLM-SLS-Pivot sheet.')

    headers = [
        str(header).strip() if header is not None and str(header).strip() else f"Column {index + 1}"
        for index, header in enumerate(values[header_index])
    ]
    metrics = headers[1:]
    if not metrics:
        raise ValueError("No target metric columns found in SLM-SLS-Pivot sheet.")

    sls_rows: list[dict[str, Any]] = []
    account_rows: list[dict[str, Any]] = []
    current_sls: dict[str, Any] | None = None
    current_group: dict[str, Any] | None = None
    current_account_rows: list[dict[str, Any]] = []
    fallback_account_rows: list[dict[str, Any]] = []

    def row_metrics(row: list[Any]) -> dict[str, float]:
        return {
            metric: _to_number(_cell(row, index + 1))
            for index, metric in enumerate(metrics)
        }

    def with_labels(metrics_payload: dict[str, float]) -> dict[str, Any]:
        return {
            "metrics": metrics_payload,
            "labels": {
                metric: _target_value_label(metric, value)
                for metric, value in metrics_payload.items()
            },
        }

    def group_fallback_row() -> dict[str, Any] | None:
        if current_sls is None or current_group is None:
            return None
        return {
            "rowNumber": current_group["rowNumber"],
            "slsName": current_sls["slsName"],
            "accountName": current_group["groupName"],
            "groupName": current_group["groupName"],
            **with_labels(current_group["metrics"]),
        }

    def finalize_current_group() -> None:
        nonlocal current_group, current_account_rows
        if current_group is None:
            return
        if current_account_rows:
            account_rows.extend(current_account_rows)
        else:
            fallback = group_fallback_row()
            if fallback is not None:
                fallback_account_rows.append(fallback)
        current_group = None
        current_account_rows = []

    def finalize_current_sls() -> None:
        finalize_current_group()
        if current_sls is not None and not _has_account_rows_for_sls(account_rows, current_sls["slsName"]):
            if fallback_account_rows:
                account_rows.extend(fallback_account_rows)
            else:
                account_rows.append(
                    {
                        "rowNumber": current_sls["rowNumber"],
                        "slsName": current_sls["slsName"],
                        "accountName": "Total",
                        "groupName": "",
                        **with_labels(current_sls["metrics"]),
                    }
                )
        fallback_account_rows.clear()

    for row_number, row in enumerate(values[header_index + 1 :], start=header_index + 2):
        label = str(_cell(row, 0) or "").strip()
        if not label:
            continue
        if label.lower() == TARGET_TOTAL_LABEL.lower():
            break

        metric_values = row_metrics(row)
        if _is_target_account_group_label(label) and current_sls is not None:
            finalize_current_group()
            current_group = {
                "rowNumber": row_number,
                "groupName": label,
                "metrics": metric_values,
            }
            continue

        if _is_target_account_label(label) and current_sls is not None:
            parent_group = current_group["groupName"] if current_group else ""
            account = (
                {
                    "rowNumber": row_number,
                    "slsName": current_sls["slsName"],
                    "accountName": _target_account_display_name(label),
                    "groupName": parent_group,
                    **with_labels(metric_values),
                }
            )
            if current_group is not None:
                current_account_rows.append(account)
            else:
                account_rows.append(account)
            continue

        finalize_current_sls()
        current_group = None
        current_sls = {
            "rowNumber": row_number,
            "slsName": label,
            **with_labels(metric_values),
        }
        sls_rows.append(current_sls)

    finalize_current_sls()
    account_rows = _target_accounts_from_data_sheet(file_bytes, filename, metrics) or account_rows

    return {
        "sheetName": sheet_name,
        "metrics": metrics,
        "slsRows": sls_rows,
        "accountRows": account_rows,
    }


def analyze_workbook(file_bytes: bytes, filename: str, sls_name: str, sheet_name: str = "Data") -> dict[str, Any]:
    _headers, rows, col_map = parse_workbook(file_bytes, filename, sheet_name)
    return analyze_forecast_rows(rows, col_map, sls_name)


def result_to_csv(result: dict[str, Any]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(["Account", "Practice", "Forecast_$K", "Target_$K", "Gap_$K"])

    for row in result.get("rows", []):
        writer.writerow(
            [
                row["account"],
                row["practice"],
                f'{row["forecast"]:.2f}',
                f'{row["target"]:.2f}',
                f'{row["gap"]:.2f}',
            ]
        )

    return output.getvalue()


def _read_sheet(file_bytes: bytes, filename: str, sheet_name: str) -> list[list[Any]]:
    extension = Path(filename or "").suffix.lower()
    if extension == ".xlsb":
        return _read_xlsb(file_bytes, sheet_name)
    if extension in {".xlsx", ".xlsm", ""}:
        return _read_xlsx(file_bytes, sheet_name)
    raise ValueError("Unsupported workbook type. Upload .xlsx, .xlsm, or .xlsb.")


def _read_xlsx(file_bytes: bytes, sheet_name: str) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found in workbook.')

    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _read_xlsb(file_bytes: bytes, sheet_name: str) -> list[list[Any]]:
    try:
        with ZipFile(io.BytesIO(file_bytes), "r") as archive:
            with XlsbWorkbook(fp=archive) as workbook:
                if sheet_name not in workbook.sheets:
                    raise ValueError(f'Sheet "{sheet_name}" not found in workbook.')

                with workbook.get_sheet(sheet_name) as sheet:
                    return [[cell.v for cell in row] for row in sheet.rows()]
    except (BadZipFile, OSError) as exc:
        raise ValueError(f"Unable to read .xlsb workbook: {exc}") from exc


def _is_target_account_group_label(value: str) -> bool:
    text = str(value or "").strip()
    return bool(re.match(r"^S\d+(?:\.\d+)?\s*[-.]", text, re.IGNORECASE))


def _is_target_account_label(value: str) -> bool:
    text = str(value or "").strip()
    return bool(re.match(r"^\d{4,}\s*-", text))


def _has_account_rows_for_sls(rows: list[dict[str, Any]], sls_name: str) -> bool:
    return any(row.get("slsName") == sls_name for row in rows)


def _target_account_display_name(value: str) -> str:
    text = str(value or "").strip()
    return re.sub(r"^\d+(?:\.\d+)?\s*-\s*", "", text).strip() or text


def _target_accounts_from_data_sheet(file_bytes: bytes, filename: str, metrics: list[str]) -> list[dict[str, Any]]:
    try:
        values = _read_sheet(file_bytes, filename, "Data")
    except ValueError:
        return []

    header_index = next(
        (
            index
            for index, row in enumerate(values)
            if "SLS" in {str(cell or "").strip() for cell in row}
            and "Parent Name" in {str(cell or "").strip() for cell in row}
        ),
        None,
    )
    if header_index is None:
        return []

    headers = [str(header or "").strip() for header in values[header_index]]
    col_map = {header: index for index, header in enumerate(headers) if header}
    required = ["SLS", "Parent Name", "SBU1"]
    if any(column not in col_map for column in required):
        return []

    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row_number, row in enumerate(values[header_index + 1 :], start=header_index + 2):
        sls_name = str(_cell(row, col_map["SLS"]) or "").strip()
        account_name = _target_account_display_name(str(_cell(row, col_map["Parent Name"]) or "").strip())
        group_name = str(_cell(row, col_map["SBU1"]) or "").strip()
        if not sls_name or not account_name:
            continue

        key = (sls_name, group_name, account_name)
        record = grouped.setdefault(
            key,
            {
                "rowNumber": row_number,
                "slsName": sls_name,
                "accountName": account_name,
                "groupName": group_name,
                "metrics": {metric: 0.0 for metric in metrics},
                "_rev_for_om": {},
                "_om_for_om": {},
            },
        )
        record["rowNumber"] = min(record["rowNumber"], row_number)

        for metric in metrics:
            if metric.startswith("OM %-"):
                suffix = metric.replace("OM %-", "", 1)
                rev_header = f"Rev-{suffix}"
                om_header = f"OM-{suffix}"
                rev_index = col_map.get(rev_header)
                om_index = col_map.get(om_header)
                record["_rev_for_om"][metric] = record["_rev_for_om"].get(metric, 0.0) + _to_number(_cell(row, rev_index) if rev_index is not None else None)
                record["_om_for_om"][metric] = record["_om_for_om"].get(metric, 0.0) + _to_number(_cell(row, om_index) if om_index is not None else None)
                continue

            data_header = metric
            if data_header in col_map:
                record["metrics"][metric] += _to_number(_cell(row, col_map[data_header]))

    account_rows: list[dict[str, Any]] = []
    for record in grouped.values():
        for metric, revenue in record.pop("_rev_for_om").items():
            om_amount = record["_om_for_om"].get(metric, 0.0)
            record["metrics"][metric] = om_amount / revenue if revenue else 0.0
        record.pop("_om_for_om", None)
        if any(abs(value) > 0.000001 for value in record["metrics"].values()):
            record["labels"] = {
                metric: _target_value_label(metric, value)
                for metric, value in record["metrics"].items()
            }
            account_rows.append(record)

    account_rows.sort(key=lambda row: (str(row["slsName"]).lower(), str(row["groupName"]).lower(), str(row["accountName"]).lower()))
    return account_rows


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None
